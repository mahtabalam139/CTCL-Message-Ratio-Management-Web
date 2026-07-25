from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from db_handler import (
    get_all_users,
    get_all_ctcl_records,
    get_all_audit_logs
)

def get_reports_page():

    return {

        "reports": [

            {
                "title": "CTCL Records Report",
                "description": "Export all CTCL records.",
                "url": "/reports/ctcl"
            },

            {
                "title": "Audit Logs Report",
                "description": "Export system audit logs.",
                "url": "/reports/audit"
            },

            {
                "title": "Users Report",
                "description": "Export all users.",
                "url": "/reports/users"
            }

        ]

    }
# ==========================================================
# EXPORT USERS REPORT
# ==========================================================

def export_users_report():

    users = get_all_users()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Users"

    sheet.append([
        "Username",
        "Full Name",
        "Role",
        "Status"
    ])
    

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for user in users:

        sheet.append(user)

    filename = "Users_Report.xlsx"

    workbook.save(filename)

    return filename

# ==========================================================
# EXPORT CTCL REPORT
# ==========================================================

def export_ctcl_report():

    records = get_all_ctcl_records()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "CTCL Records"

    headers = [

        "ID",
        "Environment",
        "Exchange",
        "Location",
        "Dedicated",
        "System",
        "Server IP",
        "Exchange IP",
        "FO CTCL",
        "CM CTCL",
        "CDS CTCL",
        "Dealer ID",
        "Rack",
        "Scenario",
        "CM Msgs",
        "FO Msgs",
        "CD Msgs",
        "Msg Line",
        "Start Date",
        "End Date",
        "Comments"

    ]

    sheet.append(headers)

    for cell in sheet[1]:

        cell.font = Font(bold=True)

    for record in records:

        sheet.append(record)

    filename = f"CTCL_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    workbook.save(filename)

    return filename

# ==========================================================
# EXPORT AUDIT REPORT
# ==========================================================

def export_audit_report():

    logs = get_all_audit_logs()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Audit Logs"

    headers = [

        "ID",
        "Username",
        "Source",
        "Module",
        "Action",
        "Description",
        "Created At"

    ]

    sheet.append(headers)

    for cell in sheet[1]:

        cell.font = Font(bold=True)

    for log in logs:

        sheet.append(log)

    filename = f"Audit_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    workbook.save(filename)

    return filename